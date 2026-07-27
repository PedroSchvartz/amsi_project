from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from database import get_db, SessionLocal
from models.lancamento import Lancamento
from models.usuario import Usuario, AcessoEnum
from models.cliente_fornecedor import ClienteFornecedor
from models.tipo_conta import tipo_conta
from schemas.lancamento import LancamentoCreate, LancamentoMassaCreate, LancamentoMassaResponse, LancamentoUpdate, LancamentoEditAdmin, LancamentoResponse, LancamentoResumo, ResumoPorTipo
from auth.dependencies import exige_admin, exige_operador_ou_admin, get_current_user
from utils.inadimplencia import atualizar_inadimplente
from utils.email_sender import enviar_email
from utils import export_jobs
from utils.config import FRONTEND_URL
from typing import List, Optional
from datetime import date, datetime, timedelta
from decimal import Decimal
import io
import re
import base64
import logging
import zipfile
import threading
from openpyxl import Workbook
from openpyxl.styles import Font

router = APIRouter(
    prefix="/lancamento",
    tags=["Lançamento"]
)


@router.get("/resumo", response_model=LancamentoResumo)
def resumo_lancamentos(
    data_pagamento_de: Optional[date] = None,
    data_pagamento_ate: Optional[date] = None,
    id_clifor: Optional[int] = None,
    id_tipo_conta: Optional[int] = None,
    natureza: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    hoje = date.today()

    def q_base():
        q = db.query(Lancamento)
        if id_clifor is not None:
            q = q.filter(Lancamento.id_clifor_relacionado_fk == id_clifor)
        if id_tipo_conta is not None:
            q = q.filter(Lancamento.id_tipo_conta_fk == id_tipo_conta)
        if natureza is not None:
            q = q.filter(Lancamento.natureza_lancamento == natureza)
        return q

    def q_periodo():
        # Dinheiro realizado: só o APROVADO conta. O que está Em análise ainda não é caixa.
        # O recorte do período continua sobre data_pagamento (a data econômica): um pagamento
        # feito em 5/jun e aprovado em 7/jun tem que contar em 5/jun, senão a fila de
        # aprovação distorce o mês de competência.
        q = q_base().filter(Lancamento.data_aprovacao != None)
        if data_pagamento_de is not None:
            q = q.filter(func.date(Lancamento.data_pagamento) >= data_pagamento_de)
        if data_pagamento_ate is not None:
            q = q.filter(func.date(Lancamento.data_pagamento) <= data_pagamento_ate)
        return q

    def soma(query, campo=None):
        col = campo if campo is not None else func.coalesce(Lancamento.valor_pago, Lancamento.valor)
        return db.query(func.coalesce(func.sum(col), 0)).filter(
            Lancamento.id_lancamento.in_(query.with_entities(Lancamento.id_lancamento))
        ).scalar()

    # Realizados no período
    q_recebido = q_periodo().filter(Lancamento.natureza_lancamento == "Credito", Lancamento.estorno == False)
    q_pago     = q_periodo().filter(Lancamento.natureza_lancamento == "Debito",  Lancamento.estorno == False)

    total_recebido = Decimal(soma(q_recebido))
    total_pago     = Decimal(soma(q_pago))

    # Reembolsos com natureza inversa: Crédito subtrai, Débito soma
    from sqlalchemy import case as sa_case
    q_reemb_ids = q_periodo().filter(Lancamento.estorno == True).with_entities(Lancamento.id_lancamento)
    total_reembolsado = db.query(
        func.coalesce(
            func.sum(
                sa_case(
                    (Lancamento.natureza_lancamento == "Credito", -func.coalesce(Lancamento.valor_pago, Lancamento.valor)),
                    else_=func.coalesce(Lancamento.valor_pago, Lancamento.valor)
                )
            ),
            0
        )
    ).filter(Lancamento.id_lancamento.in_(q_reemb_ids)).scalar()
    total_reembolsado = Decimal(total_reembolsado)

    # Saldo do período (não mais "desde sempre" — conforme spec item 5)
    saldo_total = total_recebido - total_pago

    # Pendentes: dinheiro que ainda não entrou. Em análise entra aqui — foi efetivado,
    # mas até o admin aprovar continua sendo "a receber/a pagar".
    q_abertos   = q_base().filter(Lancamento.data_aprovacao == None, Lancamento.estorno == False)
    q_a_receber = q_abertos.filter(Lancamento.natureza_lancamento == "Credito")
    q_a_pagar   = q_abertos.filter(Lancamento.natureza_lancamento == "Debito")

    total_a_receber = Decimal(db.query(func.coalesce(func.sum(Lancamento.valor), 0)).filter(
        Lancamento.id_lancamento.in_(q_a_receber.with_entities(Lancamento.id_lancamento))
    ).scalar())

    total_a_pagar = Decimal(db.query(func.coalesce(func.sum(Lancamento.valor), 0)).filter(
        Lancamento.id_lancamento.in_(q_a_pagar.with_entities(Lancamento.id_lancamento))
    ).scalar())

    # Inadimplência: créditos vencidos e não pagos
    q_vencidos_credito = q_abertos.filter(
        Lancamento.natureza_lancamento == "Credito",
        Lancamento.data_vencimento < hoje
    )
    total_inadimplencia = Decimal(db.query(func.coalesce(func.sum(Lancamento.valor), 0)).filter(
        Lancamento.id_lancamento.in_(q_vencidos_credito.with_entities(Lancamento.id_lancamento))
    ).scalar())

    # A receber excluindo inadimplentes (mantido para compatibilidade)
    ids_adimplentes = db.query(ClienteFornecedor.id_clifor).filter(ClienteFornecedor.inadimplente == False)
    total_a_receber_excl = Decimal(db.query(func.coalesce(func.sum(Lancamento.valor), 0)).filter(
        Lancamento.id_lancamento.in_(q_a_receber.with_entities(Lancamento.id_lancamento)),
        Lancamento.id_clifor_relacionado_fk.in_(ids_adimplentes)
    ).scalar())

    q_vencidos = q_abertos.filter(Lancamento.data_vencimento < hoje)
    total_vencido_a_receber = Decimal(db.query(func.coalesce(func.sum(Lancamento.valor), 0)).filter(
        Lancamento.id_lancamento.in_(q_vencidos.filter(Lancamento.natureza_lancamento == "Credito").with_entities(Lancamento.id_lancamento))
    ).scalar())
    total_vencido_a_pagar = Decimal(db.query(func.coalesce(func.sum(Lancamento.valor), 0)).filter(
        Lancamento.id_lancamento.in_(q_vencidos.filter(Lancamento.natureza_lancamento == "Debito").with_entities(Lancamento.id_lancamento))
    ).scalar())

    quantidade_abertos       = q_abertos.count()
    quantidade_vencidos      = q_vencidos.count()
    quantidade_inadimplentes = db.query(func.count(ClienteFornecedor.id_clifor)).filter(ClienteFornecedor.inadimplente == True).scalar()

    return LancamentoResumo(
        total_recebido=total_recebido,
        total_pago=total_pago,
        total_reembolsado=total_reembolsado,
        saldo_total=saldo_total,
        total_a_receber=total_a_receber,
        total_a_pagar=total_a_pagar,
        total_inadimplencia=total_inadimplencia,
        total_a_receber_excluindo_inadimplentes=total_a_receber_excl,
        total_vencido_a_receber=total_vencido_a_receber,
        total_vencido_a_pagar=total_vencido_a_pagar,
        quantidade_abertos=quantidade_abertos,
        quantidade_vencidos=quantidade_vencidos,
        quantidade_inadimplentes=quantidade_inadimplentes,
    )


@router.get("/resumo-por-tipo", response_model=List[ResumoPorTipo])
def resumo_por_tipo(
    data_pagamento_de: Optional[date] = None,
    data_pagamento_ate: Optional[date] = None,
    natureza: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    q = (
        db.query(
            tipo_conta.id_tipo_conta,
            tipo_conta.descricao_conta,
            tipo_conta.natureza_conta,
            func.coalesce(func.sum(func.coalesce(Lancamento.valor_pago, Lancamento.valor)), 0).label("total"),
            func.count(Lancamento.id_lancamento).label("quantidade")
        )
        .join(Lancamento, Lancamento.id_tipo_conta_fk == tipo_conta.id_tipo_conta)
        .filter(Lancamento.data_aprovacao != None, Lancamento.estorno == False)   # dinheiro: só aprovado
    )
    if data_pagamento_de is not None:
        q = q.filter(func.date(Lancamento.data_pagamento) >= data_pagamento_de)
    if data_pagamento_ate is not None:
        q = q.filter(func.date(Lancamento.data_pagamento) <= data_pagamento_ate)
    if natureza is not None:
        q = q.filter(tipo_conta.natureza_conta == natureza)

    q = q.group_by(tipo_conta.id_tipo_conta, tipo_conta.descricao_conta, tipo_conta.natureza_conta)
    q = q.order_by(func.sum(func.coalesce(Lancamento.valor_pago, Lancamento.valor)).desc())

    return [
        ResumoPorTipo(
            id_tipo_conta=r.id_tipo_conta,
            descricao_conta=r.descricao_conta,
            natureza_conta=r.natureza_conta,
            total=Decimal(r.total),
            quantidade=r.quantidade
        )
        for r in q.all()
    ]


def _query_lancamentos_filtrada(
    db: Session,
    id_clifor: Optional[int] = None,
    id_tipo_conta: Optional[int] = None,
    natureza: Optional[str] = None,
    apenas_abertos: Optional[bool] = None,
    apenas_vencidos: Optional[bool] = None,
    apenas_em_analise: Optional[bool] = None,
    apenas_quitados: Optional[bool] = None,
    apenas_com_comprovante: Optional[bool] = None,
    apenas_sem_comprovante: Optional[bool] = None,
    data_vencimento_de: Optional[date] = None,
    data_vencimento_ate: Optional[date] = None,
    data_lancamento_de: Optional[date] = None,
    data_lancamento_ate: Optional[date] = None,
    data_pagamento_de: Optional[date] = None,
    data_pagamento_ate: Optional[date] = None,
    estorno: Optional[bool] = None,
    valor_minimo: Optional[Decimal] = None,
    valor_maximo: Optional[Decimal] = None,
    lote: Optional[int] = None,
):
    """Monta a query filtrada da listagem. Compartilhada entre GET / (JSON) e a
    exportação em .xlsx — assim os dois nunca divergem nos filtros aplicados."""
    query = db.query(Lancamento).join(
        ClienteFornecedor,
        Lancamento.id_clifor_relacionado_fk == ClienteFornecedor.id_clifor
    ).options(
        joinedload(Lancamento.cliente_fornecedor),
        joinedload(Lancamento.usuario_lancamento),
        joinedload(Lancamento.usuario_efetivacao),
        joinedload(Lancamento.usuario_aprovacao),
        joinedload(Lancamento.usuario_edicao),
        joinedload(Lancamento.tipo_conta_rel)
    )

    if id_clifor is not None:
        query = query.filter(Lancamento.id_clifor_relacionado_fk == id_clifor)
    if id_tipo_conta is not None:
        query = query.filter(Lancamento.id_tipo_conta_fk == id_tipo_conta)
    if natureza is not None:
        query = query.filter(Lancamento.natureza_lancamento == natureza)
    # Estes filtros são de ESTADO (o que o usuário vê na tela), não de dinheiro:
    # olham data_efetivacao para casar exatamente com o badge da listagem.
    if apenas_abertos:
        query = query.filter(Lancamento.data_efetivacao == None)
    if apenas_vencidos:
        # Em análise não é vencido: alguém já pagou, não se cobra de novo.
        query = query.filter(
            Lancamento.data_efetivacao == None,
            Lancamento.data_vencimento < date.today()
        )
    if apenas_em_analise:
        query = query.filter(Lancamento.data_efetivacao != None, Lancamento.data_aprovacao == None)
    if apenas_quitados:
        query = query.filter(Lancamento.data_aprovacao != None)
    if apenas_com_comprovante:
        query = query.filter(Lancamento.comprovante != None)
    if apenas_sem_comprovante:
        query = query.filter(Lancamento.comprovante == None)
    if data_vencimento_de is not None:
        query = query.filter(Lancamento.data_vencimento >= data_vencimento_de)
    if data_vencimento_ate is not None:
        query = query.filter(Lancamento.data_vencimento <= data_vencimento_ate)
    if data_lancamento_de is not None:
        query = query.filter(Lancamento.data_lancamento >= data_lancamento_de)
    if data_lancamento_ate is not None:
        query = query.filter(Lancamento.data_lancamento <= data_lancamento_ate)
    if data_pagamento_de is not None:
        query = query.filter(func.date(Lancamento.data_pagamento) >= data_pagamento_de)
    if data_pagamento_ate is not None:
        query = query.filter(func.date(Lancamento.data_pagamento) <= data_pagamento_ate)
    if estorno is not None:
        query = query.filter(Lancamento.estorno == estorno)
    if valor_minimo is not None:
        query = query.filter(Lancamento.valor >= valor_minimo)
    if valor_maximo is not None:
        query = query.filter(Lancamento.valor <= valor_maximo)
    if lote is not None:
        query = query.filter(Lancamento.lote == lote)

    return query.order_by(Lancamento.data_vencimento, ClienteFornecedor.nome)


@router.get("/", response_model=List[LancamentoResponse])
def listar_lancamentos(
    id_clifor: Optional[int] = None,
    id_tipo_conta: Optional[int] = None,
    natureza: Optional[str] = None,
    apenas_abertos: Optional[bool] = None,
    apenas_vencidos: Optional[bool] = None,
    apenas_em_analise: Optional[bool] = None,
    apenas_quitados: Optional[bool] = None,
    apenas_com_comprovante: Optional[bool] = None,
    apenas_sem_comprovante: Optional[bool] = None,
    data_vencimento_de: Optional[date] = None,
    data_vencimento_ate: Optional[date] = None,
    data_lancamento_de: Optional[date] = None,
    data_lancamento_ate: Optional[date] = None,
    data_pagamento_de: Optional[date] = None,
    data_pagamento_ate: Optional[date] = None,
    estorno: Optional[bool] = None,
    valor_minimo: Optional[Decimal] = None,
    valor_maximo: Optional[Decimal] = None,
    lote: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    return _query_lancamentos_filtrada(
        db,
        id_clifor=id_clifor, id_tipo_conta=id_tipo_conta, natureza=natureza,
        apenas_abertos=apenas_abertos, apenas_vencidos=apenas_vencidos,
        apenas_em_analise=apenas_em_analise, apenas_quitados=apenas_quitados,
        apenas_com_comprovante=apenas_com_comprovante, apenas_sem_comprovante=apenas_sem_comprovante,
        data_vencimento_de=data_vencimento_de, data_vencimento_ate=data_vencimento_ate,
        data_lancamento_de=data_lancamento_de, data_lancamento_ate=data_lancamento_ate,
        data_pagamento_de=data_pagamento_de, data_pagamento_ate=data_pagamento_ate,
        estorno=estorno, valor_minimo=valor_minimo, valor_maximo=valor_maximo, lote=lote,
    ).all()


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTAÇÃO EM .XLSX — assíncrona (job em memória; ver utils/export_jobs.py)
# ══════════════════════════════════════════════════════════════════════════════

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Formatos de data: campos de data pura / digitados pelo usuário NÃO levam fuso
# (data_vencimento é DATE; data_pagamento é a data digitada à meia-noite). Só os
# carimbos do servidor (UTC) viram horário local — Brasil hoje é UTC-3 fixo, sem
# horário de verão, então subtrair 3h basta e não exige tzdata.
def _fmt_data(d) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def _fmt_carimbo(dt) -> str:
    return (dt - timedelta(hours=3)).strftime("%d/%m/%Y %H:%M") if dt else ""


def _num(v):
    return float(v) if v is not None else None


def _mascarar_cpf_cnpj(doc: Optional[str]) -> str:
    """Espelha rassurarCpfCnpj do frontend — máscara para o perfil Consulta."""
    if not doc:
        return "—"
    d = re.sub(r"\D", "", doc)
    if len(d) == 11:
        return f"***.***.{d[6:9]}-**"
    if len(d) == 14:
        return f"**.{d[2:5]}.{d[5:8]}/****.{d[12:]}"
    return doc


_NATUREZA_LABEL = {"Debito": "Débito", "Credito": "Crédito"}


def _valor_enum(v) -> str:
    return getattr(v, "value", v) or ""


def _total_pago(r) -> Optional[float]:
    if r.valor_pago is None:
        return None
    return float(r.valor_pago or 0) + float(r.multa or 0) + float(r.juros or 0)


def _descrever_filtros(db: Session, f: dict) -> str:
    """Texto legível dos filtros — vai no topo do .xlsx e no corpo do e-mail."""
    partes = []
    if f.get("id_clifor") is not None:
        cf = db.query(ClienteFornecedor).filter(ClienteFornecedor.id_clifor == f["id_clifor"]).first()
        partes.append(f"Cliente/Fornecedor: {cf.nome if cf else f['id_clifor']}")
    if f.get("id_tipo_conta") is not None:
        tc = db.query(tipo_conta).filter(tipo_conta.id_tipo_conta == f["id_tipo_conta"]).first()
        partes.append(f"Tipo de Conta: {tc.descricao_conta if tc else f['id_tipo_conta']}")
    if f.get("natureza"):
        partes.append(f"Natureza: {_NATUREZA_LABEL.get(f['natureza'], f['natureza'])}")
    status = [rot for chave, rot in (
        ("apenas_abertos", "Abertos"), ("apenas_vencidos", "Vencidos"),
        ("apenas_em_analise", "Em análise"), ("apenas_quitados", "Quitados"),
    ) if f.get(chave)]
    if status:
        partes.append(f"Status: {', '.join(status)}")
    if f.get("apenas_com_comprovante"):
        partes.append("Com comprovante")
    if f.get("apenas_sem_comprovante"):
        partes.append("Sem comprovante")
    if f.get("estorno") is not None:
        partes.append(f"Reembolso: {'sim' if f['estorno'] else 'não'}")
    if f.get("data_vencimento_de") or f.get("data_vencimento_ate"):
        partes.append(f"Vencimento: {_fmt_data(f.get('data_vencimento_de')) or '…'} a {_fmt_data(f.get('data_vencimento_ate')) or '…'}")
    if f.get("data_lancamento_de") or f.get("data_lancamento_ate"):
        partes.append(f"Lançamento: {_fmt_data(f.get('data_lancamento_de')) or '…'} a {_fmt_data(f.get('data_lancamento_ate')) or '…'}")
    if f.get("data_pagamento_de") or f.get("data_pagamento_ate"):
        partes.append(f"Pagamento: {_fmt_data(f.get('data_pagamento_de')) or '…'} a {_fmt_data(f.get('data_pagamento_ate')) or '…'}")
    if f.get("valor_minimo") is not None or f.get("valor_maximo") is not None:
        partes.append(f"Valor: {f.get('valor_minimo', '…')} a {f.get('valor_maximo', '…')}")
    return " · ".join(partes) if partes else "Nenhum filtro aplicado (todos os lançamentos)"


def _colunas(r, perfil_completo: bool) -> list[tuple[str, object]]:
    """Pares (cabeçalho, valor) de uma linha. Consulta recebe só as colunas da
    tabela (CPF mascarado); Operador/Admin recebem tabela + campos das modais."""
    origem = "Em Lote" if r.lote is not None else "Manual"
    tipo = f"{r.id_tipo_conta_fk} - {r.descricao_tipo_conta}" if r.descricao_tipo_conta else str(r.id_tipo_conta_fk)
    natureza = _NATUREZA_LABEL.get(_valor_enum(r.natureza_lancamento), _valor_enum(r.natureza_lancamento))

    if not perfil_completo:
        return [
            ("CPF/CNPJ", _mascarar_cpf_cnpj(r.cpf_cnpj_clifor)),
            ("Nome / Razão Social", r.nome_clifor or "—"),
            ("Tipo de Conta", tipo),
            ("Natureza", natureza),
            ("Vencimento", _fmt_data(r.data_vencimento)),
            ("Pagamento", _fmt_data(r.data_pagamento)),
            ("Vl. Lançamento", _num(r.valor)),
            ("Vl. Pagamento", _total_pago(r)),
            ("Status", _valor_enum(r.situacao)),
            ("Origem", origem),
        ]

    return [
        ("ID", r.id_lancamento),
        ("CPF/CNPJ", r.cpf_cnpj_clifor or "—"),
        ("Nome / Razão Social", r.nome_clifor or "—"),
        ("Tipo de Conta", tipo),
        ("Natureza", natureza),
        ("Status", _valor_enum(r.situacao)),
        ("Origem", origem),
        ("Lote", r.lote if r.lote is not None else ""),
        ("Reembolso", "Sim" if r.estorno else "Não"),
        ("Data de Lançamento", _fmt_carimbo(r.data_lancamento)),
        ("Vencimento", _fmt_data(r.data_vencimento)),
        ("Pagamento", _fmt_data(r.data_pagamento)),
        ("Vl. Lançamento", _num(r.valor)),
        ("Valor Pago", _num(r.valor_pago)),
        ("Multa", _num(r.multa)),
        ("Juros", _num(r.juros)),
        ("Vl. Pagamento (Total)", _total_pago(r)),
        ("Observação do Lançamento", r.observacao or ""),
        ("Observação do Pagamento", r.observacao_pagamento or ""),
        ("Comprovante", r.comprovante_nome or "—"),
        ("Lançado por", r.nome_usuario_lancamento or "—"),
        ("Efetivado por", r.nome_usuario_efetivacao or "—"),
        ("Data de Efetivação", _fmt_carimbo(r.data_efetivacao)),
        ("Aprovado por", r.nome_usuario_aprovacao or "—"),
        ("Data de Aprovação", _fmt_carimbo(r.data_aprovacao)),
        ("Editado por", r.nome_usuario_edicao or "—"),
        ("Data de Edição", _fmt_carimbo(r.data_edicao)),
    ]


def _montar_xlsx(registros, perfil_completo: bool, filtros_desc: str, data_pesquisa: datetime, job_id: str) -> Optional[bytes]:
    """Monta o workbook. Retorna None se o job foi cancelado no meio."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    # Bloco de metadados — também vai no corpo do e-mail.
    ws["A1"] = "Exportação de Lançamentos — AMSI"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Data da pesquisa: {_fmt_carimbo(data_pesquisa)}"
    ws["A3"] = f"Filtros: {filtros_desc}"
    ws["A4"] = f"Total de lançamentos: {len(registros)}"

    LINHA_CABECALHO = 6
    cabecalhos = [h for h, _ in _colunas(registros[0], perfil_completo)] if registros else \
        [h for h, _ in _colunas_vazia(perfil_completo)]
    for col, titulo in enumerate(cabecalhos, start=1):
        c = ws.cell(row=LINHA_CABECALHO, column=col, value=titulo)
        c.font = Font(bold=True)
    ws.freeze_panes = f"A{LINHA_CABECALHO + 1}"

    larguras = [len(h) for h in cabecalhos]
    linha = LINHA_CABECALHO + 1
    for i, r in enumerate(registros):
        if i % 200 == 0 and export_jobs.cancelado(job_id):
            return None
        for col, (_, valor) in enumerate(_colunas(r, perfil_completo), start=1):
            ws.cell(row=linha, column=col, value=valor)
            larguras[col - 1] = max(larguras[col - 1], len(str(valor)) if valor is not None else 0)
        linha += 1

    # Largura de coluna aproximada, limitada para não estourar.
    from openpyxl.utils import get_column_letter
    for col, larg in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(larg + 2, 10), 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _colunas_vazia(perfil_completo: bool):
    """Cabeçalhos quando não há registros — usa um objeto stub só para os títulos."""
    class _Stub:
        def __getattr__(self, _):
            return None
    return _colunas(_Stub(), perfil_completo)


def _processar_exportacao(job_id: str, filtros: dict, perfil_completo: bool,
                          filtros_desc: str, data_pesquisa: datetime):
    """Roda em thread própria: consulta o banco e monta o .xlsx, guardando no job."""
    db = SessionLocal()
    try:
        registros = [
            LancamentoResponse.model_validate(l)
            for l in _query_lancamentos_filtrada(db, **filtros).all()
        ]
    except Exception as e:
        logging.warning(f"Exportação {job_id}: falha na consulta — {e}")
        export_jobs.marcar_erro(job_id, "Falha ao consultar os lançamentos.")
        db.close()
        return
    finally:
        db.close()

    if export_jobs.cancelado(job_id):
        return
    try:
        conteudo = _montar_xlsx(registros, perfil_completo, filtros_desc, data_pesquisa, job_id)
        if conteudo is None:  # cancelado no meio da montagem
            return
        # Horário local (UTC-3) no nome, com data e hora:min:seg.
        data_str = (data_pesquisa - timedelta(hours=3)).strftime("%Y-%m-%d_%H-%M-%S")
        export_jobs.marcar_concluido(job_id, conteudo, f"lancamentos_{data_str}.xlsx")
    except Exception as e:
        logging.warning(f"Exportação {job_id}: falha ao montar planilha — {e}")
        export_jobs.marcar_erro(job_id, "Falha ao gerar a planilha.")


def _dono_ou_404(job_id: str, current_user: Usuario) -> dict:
    """Recupera o job garantindo que pertence ao usuário — o arquivo pode conter
    CPF sem máscara, então um usuário nunca acessa job de outro."""
    job = export_jobs.obter(job_id)
    if not job or job["user_id"] != current_user.id_usuario:
        raise HTTPException(status_code=404, detail="Exportação não encontrada")
    return job


@router.post("/exportar")
def iniciar_exportacao(
    id_clifor: Optional[int] = None,
    id_tipo_conta: Optional[int] = None,
    natureza: Optional[str] = None,
    apenas_abertos: Optional[bool] = None,
    apenas_vencidos: Optional[bool] = None,
    apenas_em_analise: Optional[bool] = None,
    apenas_quitados: Optional[bool] = None,
    apenas_com_comprovante: Optional[bool] = None,
    apenas_sem_comprovante: Optional[bool] = None,
    data_vencimento_de: Optional[date] = None,
    data_vencimento_ate: Optional[date] = None,
    data_lancamento_de: Optional[date] = None,
    data_lancamento_ate: Optional[date] = None,
    data_pagamento_de: Optional[date] = None,
    data_pagamento_ate: Optional[date] = None,
    estorno: Optional[bool] = None,
    valor_minimo: Optional[Decimal] = None,
    valor_maximo: Optional[Decimal] = None,
    lote: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    """Inicia a geração de um .xlsx com os lançamentos filtrados (nova consulta ao
    banco). Devolve um job_id para acompanhamento; a entrega (download/e-mail) é
    decidida pelas rotas seguintes."""
    filtros = {
        "id_clifor": id_clifor, "id_tipo_conta": id_tipo_conta, "natureza": natureza,
        "apenas_abertos": apenas_abertos, "apenas_vencidos": apenas_vencidos,
        "apenas_em_analise": apenas_em_analise, "apenas_quitados": apenas_quitados,
        "apenas_com_comprovante": apenas_com_comprovante, "apenas_sem_comprovante": apenas_sem_comprovante,
        "data_vencimento_de": data_vencimento_de, "data_vencimento_ate": data_vencimento_ate,
        "data_lancamento_de": data_lancamento_de, "data_lancamento_ate": data_lancamento_ate,
        "data_pagamento_de": data_pagamento_de, "data_pagamento_ate": data_pagamento_ate,
        "estorno": estorno, "valor_minimo": valor_minimo, "valor_maximo": valor_maximo, "lote": lote,
    }
    data_pesquisa = datetime.utcnow()
    filtros_desc = _descrever_filtros(db, filtros)
    perfil_completo = current_user.perfil_de_acesso in (AcessoEnum.Administrador, AcessoEnum.Operador)

    job_id = export_jobs.criar(current_user.id_usuario, filtros_desc, data_pesquisa)
    threading.Thread(
        target=_processar_exportacao,
        args=(job_id, filtros, perfil_completo, filtros_desc, data_pesquisa),
        daemon=True,
    ).start()
    return {"job_id": job_id}


@router.get("/exportar/{job_id}")
def status_exportacao(job_id: str, current_user: Usuario = Depends(get_current_user)):
    job = _dono_ou_404(job_id, current_user)
    return {"status": job["status"], "error": job["error"]}


@router.get("/exportar/{job_id}/download")
def baixar_exportacao(job_id: str, current_user: Usuario = Depends(get_current_user)):
    job = _dono_ou_404(job_id, current_user)
    if job["status"] != export_jobs.CONCLUIDO or not job["resultado"]:
        raise HTTPException(status_code=409, detail="Exportação ainda não está pronta")
    return Response(
        content=job["resultado"],
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{job["filename"]}"'},
    )


@router.post("/exportar/{job_id}/enviar-email")
def enviar_exportacao_email(job_id: str, current_user: Usuario = Depends(get_current_user)):
    job = _dono_ou_404(job_id, current_user)
    if job["status"] != export_jobs.CONCLUIDO or not job["resultado"]:
        raise HTTPException(status_code=409, detail="Exportação ainda não está pronta")

    # Zipa o .xlsx e anexa em base64 (formato da Brevo).
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(job["filename"], job["resultado"])
    zip_nome = job["filename"].rsplit(".", 1)[0] + ".zip"
    anexo_b64 = base64.b64encode(zip_buffer.getvalue()).decode("ascii")

    corpo = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<body style="margin:0;padding:0;background:#EFE6DD;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="padding:40px 20px;">
    <tr><td align="center">
      <table width="600" style="background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(27,67,50,0.10);">
        <tr><td style="background:#1B4332;padding:32px 40px;text-align:center;">
          <p style="margin:0;font-size:2rem;font-weight:700;color:#C9A84C;letter-spacing:0.1em;">AMSI</p>
          <p style="margin:4px 0 0;font-size:0.72rem;color:rgba(255,255,255,0.6);letter-spacing:0.2em;text-transform:uppercase;">Associação de Moradores de Santa Isabel</p>
        </td></tr>
        <tr><td style="padding:36px 40px;">
          <p style="font-size:1.3rem;font-weight:600;color:#1B4332;margin:0 0 8px;">Exportação de lançamentos 📊</p>
          <p style="color:#6b7280;margin:0 0 20px;">Olá, <strong style="color:#2C2C2C;">{current_user.nome}</strong>! Segue em anexo (arquivo <strong>.zip</strong>) o resultado da sua exportação de lançamentos.</p>
          <div style="background:#f4f1ec;border-left:4px solid #1B4332;padding:12px 16px;border-radius:4px;margin:0 0 20px;">
            <p style="margin:0 0 6px;font-size:0.85rem;color:#2C2C2C;"><strong>Data da pesquisa:</strong> {_fmt_carimbo(job["data_pesquisa"])}</p>
            <p style="margin:0;font-size:0.85rem;color:#2C2C2C;"><strong>Filtros aplicados:</strong> {job["filtros_desc"]}</p>
          </div>
          <p style="font-size:0.78rem;color:#6b7280;margin:0;">Se preferir, acesse o sistema: <a href="{FRONTEND_URL}" style="color:#1B4332;">{FRONTEND_URL}</a></p>
        </td></tr>
        <tr><td style="padding:16px 40px;text-align:center;border-top:1px solid #d1c9bf;">
          <p style="margin:0;font-size:0.72rem;color:#a0a0a0;">© 2026 AMSI — Este é um email automático.</p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body>
</html>
"""
    enviado = enviar_email(
        current_user.email,
        "Exportação de lançamentos — AMSI Project",
        corpo,
        anexos=[{"content": anexo_b64, "name": zip_nome}],
    )
    if not enviado:
        raise HTTPException(status_code=502, detail="Não foi possível enviar o e-mail. Tente novamente.")
    return {"detail": "E-mail enviado com sucesso"}


@router.delete("/exportar/{job_id}")
def cancelar_exportacao(job_id: str, current_user: Usuario = Depends(get_current_user)):
    _dono_ou_404(job_id, current_user)
    export_jobs.cancelar(job_id)
    export_jobs.marcar_cancelado(job_id)
    return {"detail": "Exportação cancelada"}


@router.get("/{id_lancamento}", response_model=LancamentoResponse)
def buscar_lancamento(id_lancamento: int, db: Session = Depends(get_db), _=Depends(exige_operador_ou_admin)):
    lancamento = db.query(Lancamento).filter(Lancamento.id_lancamento == id_lancamento).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    return lancamento


@router.get("/por-clifor/{id_clifor}", response_model=List[LancamentoResponse])
def listar_lancamentos_por_clifor(id_clifor: int, db: Session = Depends(get_db), _=Depends(exige_operador_ou_admin)):
    return db.query(Lancamento).filter(
        Lancamento.id_clifor_relacionado_fk == id_clifor
    ).options(
        joinedload(Lancamento.cliente_fornecedor),
        joinedload(Lancamento.usuario_lancamento),
        joinedload(Lancamento.usuario_efetivacao),
        joinedload(Lancamento.usuario_aprovacao),
        joinedload(Lancamento.usuario_edicao),
        joinedload(Lancamento.tipo_conta_rel)
    ).order_by(Lancamento.data_vencimento).all()


@router.get("/por-usuario/{id_usuario}", response_model=List[LancamentoResponse])
def listar_lancamentos_por_usuario(id_usuario: int, db: Session = Depends(get_db), _=Depends(exige_operador_ou_admin)):
    return db.query(Lancamento).filter(
        Lancamento.id_usuario_fk_lancamento == id_usuario
    ).join(
        ClienteFornecedor,
        Lancamento.id_clifor_relacionado_fk == ClienteFornecedor.id_clifor
    ).options(
        joinedload(Lancamento.usuario_lancamento),
        joinedload(Lancamento.usuario_efetivacao),
        joinedload(Lancamento.usuario_aprovacao),
        joinedload(Lancamento.usuario_edicao),
        joinedload(Lancamento.tipo_conta_rel)
    ).order_by(Lancamento.data_vencimento, ClienteFornecedor.nome).all()


def _montar_lancamento(dados: dict) -> Lancamento:
    """Monta a ORM Lancamento a partir de um dict — sem validar nem commitar.

    Compartilhado entre a criação única e a criação em massa para que as duas rotas
    não divirjam na construção do objeto.
    """
    return Lancamento(**dados)


@router.post("/", response_model=LancamentoResponse)
def criar_lancamento(dados: LancamentoCreate, db: Session = Depends(get_db), _=Depends(exige_operador_ou_admin)):
    if not db.query(Usuario).filter(Usuario.id_usuario == dados.id_usuario_fk_lancamento).first():
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    if not db.query(ClienteFornecedor).filter(ClienteFornecedor.id_clifor == dados.id_clifor_relacionado_fk).first():
        raise HTTPException(status_code=404, detail="Cliente/Fornecedor não encontrado")
    if not db.query(tipo_conta).filter(tipo_conta.id_tipo_conta == dados.id_tipo_conta_fk).first():
        raise HTTPException(status_code=404, detail="Tipo de lançamento não encontrado")
    lancamento = _montar_lancamento(dados.model_dump())
    db.add(lancamento)
    db.commit()
    db.refresh(lancamento)
    atualizar_inadimplente(lancamento.id_clifor_relacionado_fk, db)
    return lancamento


@router.post("/massa", response_model=LancamentoMassaResponse)
def criar_lancamentos_massa(
    dados: LancamentoMassaCreate,
    db: Session = Depends(get_db),
    _=Depends(exige_operador_ou_admin),
):
    if not db.query(Usuario).filter(Usuario.id_usuario == dados.id_usuario_fk_lancamento).first():
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    if not db.query(tipo_conta).filter(tipo_conta.id_tipo_conta == dados.id_tipo_conta_fk).first():
        raise HTTPException(status_code=404, detail="Tipo de lançamento não encontrado")

    ids = list(dict.fromkeys(dados.ids_clifor))   # dedup preservando ordem
    if not ids:
        raise HTTPException(status_code=422, detail="Selecione ao menos um cliente/fornecedor")

    encontrados = {
        r.id_clifor for r in
        db.query(ClienteFornecedor.id_clifor).filter(ClienteFornecedor.id_clifor.in_(ids)).all()
    }
    faltantes = [i for i in ids if i not in encontrados]
    if faltantes:
        raise HTTPException(status_code=404, detail=f"Cliente/Fornecedor não encontrado: {faltantes}")

    lote = int(datetime.utcnow().timestamp() * 1000)   # ms, igual ao exp_ms do token

    novos = []
    for id_clifor in ids:
        l = _montar_lancamento({
            "id_usuario_fk_lancamento": dados.id_usuario_fk_lancamento,
            "id_clifor_relacionado_fk": id_clifor,
            "id_tipo_conta_fk": dados.id_tipo_conta_fk,
            "valor": dados.valor,
            "data_vencimento": dados.data_vencimento,
            "natureza_lancamento": dados.natureza_lancamento,
            "observacao": dados.observacao,
            "lote": lote,
        })
        db.add(l)
        novos.append(l)

    db.commit()                                  # atômico: 1 commit para todo o lote
    for l in novos:
        db.refresh(l)
    for id_clifor in set(ids):
        atualizar_inadimplente(id_clifor, db)

    return LancamentoMassaResponse(
        lote=lote,
        total_criados=len(novos),
        ids=[l.id_lancamento for l in novos],
    )


@router.put("/{id_lancamento}", response_model=LancamentoResponse)
def efetivar_lancamento(
    id_lancamento: int,
    dados: LancamentoUpdate,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(exige_operador_ou_admin)
):
    """Efetiva um lançamento: Aberto → Em análise.

    É uma TRANSIÇÃO, não um update solto — daí o 409 em cima de quem já foi efetivado.
    Sem essa checagem o portão seria contornável: bastaria ao Operador dar um segundo PUT
    depois da aprovação para trocar valor_pago sem passar por análise nenhuma.
    Correção depois de efetivado é com o admin, pelo PATCH /editar.

    Admin vai direto para Pago: ele aprovaria em seguida de qualquer jeito, então acumula
    os dois papéis na mesma operação.
    """
    lancamento = db.query(Lancamento).filter(Lancamento.id_lancamento == id_lancamento).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    if lancamento.data_efetivacao is not None:
        raise HTTPException(status_code=409, detail="Lançamento já efetivado")
    if not dados.data_pagamento:
        raise HTTPException(status_code=422, detail="data_pagamento é obrigatório para efetivar um lançamento")

    for campo, valor in dados.model_dump(exclude_unset=True).items():
        setattr(lancamento, campo, valor)

    # Atores e carimbos saem do token — o corpo não opina. É o que faz o registro de
    # quem mandou para análise valer alguma coisa.
    agora = datetime.utcnow()
    lancamento.data_efetivacao = agora
    lancamento.id_usuario_fk_efetivacao = current_user.id_usuario
    if current_user.perfil_de_acesso == AcessoEnum.Administrador:
        lancamento.data_aprovacao = agora
        lancamento.id_usuario_fk_aprovacao = current_user.id_usuario

    db.commit()
    db.refresh(lancamento)
    atualizar_inadimplente(lancamento.id_clifor_relacionado_fk, db)
    return lancamento


@router.post("/{id_lancamento}/aprovar", response_model=LancamentoResponse)
def aprovar_lancamento(
    id_lancamento: int,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(exige_admin)
):
    """Aprova um lançamento: Em análise → Pago. Só administrador.

    É a aprovação que faz o dinheiro entrar no caixa e limpa a inadimplência —
    até aqui, o lançamento efetivado não conta em resumo, dashboard nem saldo.
    """
    lancamento = db.query(Lancamento).filter(Lancamento.id_lancamento == id_lancamento).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    if lancamento.data_efetivacao is None:
        raise HTTPException(status_code=409, detail="Lançamento não está em análise: efetive antes de aprovar")
    if lancamento.data_aprovacao is not None:
        raise HTTPException(status_code=409, detail="Lançamento já aprovado")

    lancamento.data_aprovacao = datetime.utcnow()
    lancamento.id_usuario_fk_aprovacao = admin.id_usuario

    db.commit()
    db.refresh(lancamento)
    atualizar_inadimplente(lancamento.id_clifor_relacionado_fk, db)
    return lancamento


@router.patch("/{id_lancamento}/editar", response_model=LancamentoResponse)
def editar_lancamento_admin(
    id_lancamento: int,
    dados: LancamentoEditAdmin,
    db: Session = Depends(get_db),
    admin=Depends(exige_admin)
):
    """Edição completa de um lançamento — restrita a administradores."""
    lancamento = db.query(Lancamento).filter(Lancamento.id_lancamento == id_lancamento).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    if dados.id_clifor_relacionado_fk is not None:
        if not db.query(ClienteFornecedor).filter(ClienteFornecedor.id_clifor == dados.id_clifor_relacionado_fk).first():
            raise HTTPException(status_code=404, detail="Cliente/Fornecedor não encontrado")
    if dados.id_tipo_conta_fk is not None:
        if not db.query(tipo_conta).filter(tipo_conta.id_tipo_conta == dados.id_tipo_conta_fk).first():
            raise HTTPException(status_code=404, detail="Tipo de conta não encontrado")
    for campo, valor in dados.model_dump(exclude_unset=True).items():
        setattr(lancamento, campo, valor)

    # Invariante do fluxo, conferida sobre o objeto já mesclado: como é um PATCH parcial,
    # o payload sozinho não diz o estado final. data_efetivacao não é editável, mas
    # data_pagamento é — e zerá-la num lançamento já efetivado quebraria o par.
    # Espelha o CHECK do banco; sem isto o erro viria como 500 do Postgres.
    if lancamento.data_efetivacao is not None and lancamento.data_pagamento is None:
        raise HTTPException(status_code=400, detail="Lançamento efetivado exige data_pagamento")

    # Carimbo da edição — do token, nunca do corpo. Sobrescreve a edição anterior:
    # a linha do tempo guarda só a última.
    lancamento.data_edicao = datetime.utcnow()
    lancamento.id_usuario_fk_edicao = admin.id_usuario

    db.commit()
    db.refresh(lancamento)
    atualizar_inadimplente(lancamento.id_clifor_relacionado_fk, db)
    return lancamento


@router.post("/{id_lancamento}/comprovante")
async def anexar_comprovante(
    id_lancamento: int,
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(exige_operador_ou_admin)
):
    lancamento = db.query(Lancamento).filter(Lancamento.id_lancamento == id_lancamento).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")

    if arquivo.content_type != "application/pdf":
        raise HTTPException(status_code=400, detail="Apenas arquivos PDF são aceitos")

    conteudo = await arquivo.read()
    if len(conteudo) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Arquivo muito grande. Máximo 5MB.")

    lancamento.comprovante = conteudo
    lancamento.comprovante_nome = arquivo.filename
    db.commit()
    return {"detail": "Comprovante anexado com sucesso", "nome": arquivo.filename}


@router.get("/{id_lancamento}/comprovante")
def baixar_comprovante(
    id_lancamento: int,
    db: Session = Depends(get_db),
    _=Depends(exige_operador_ou_admin)
):
    lancamento = db.query(Lancamento).filter(Lancamento.id_lancamento == id_lancamento).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    if not lancamento.comprovante:
        raise HTTPException(status_code=404, detail="Este lançamento não possui comprovante")

    return Response(
        content=lancamento.comprovante,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{lancamento.comprovante_nome or "comprovante.pdf"}"'}
    )


@router.delete("/{id_lancamento}/comprovante")
def remover_comprovante(
    id_lancamento: int,
    db: Session = Depends(get_db),
    _=Depends(exige_operador_ou_admin)
):
    lancamento = db.query(Lancamento).filter(Lancamento.id_lancamento == id_lancamento).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    if not lancamento.comprovante:
        raise HTTPException(status_code=404, detail="Este lançamento não possui comprovante")
    lancamento.comprovante = None
    lancamento.comprovante_nome = None
    db.commit()
    return {"detail": "Comprovante removido com sucesso"}


@router.delete("/{id_lancamento}")
def deletar_lancamento(id_lancamento: int, db: Session = Depends(get_db), _=Depends(exige_admin)):
    lancamento = db.query(Lancamento).filter(Lancamento.id_lancamento == id_lancamento).first()
    if not lancamento:
        raise HTTPException(status_code=404, detail="Lançamento não encontrado")
    id_clifor = lancamento.id_clifor_relacionado_fk
    db.delete(lancamento)
    db.commit()
    atualizar_inadimplente(id_clifor, db)
    return {"mensagem": "Lançamento deletado com sucesso"}