"""
Script de importação de Clientes/Fornecedores a partir de XLSX.

Fluxo por linha:
  1. Valida campos obrigatórios (CPF, RG, Nome Completo + Celular ou E-mail)
  2. Cria o CliFor  → captura id_clifor
  3. Cria contato(s) — Celular e/ou E-mail
  4. Se endereço completo, cria o endereço vinculado
  5. Linhas inválidas vão para 'pendentes_<arquivo_original>.xlsx'
     com uma coluna extra "Campos Faltando" indicando o que está ausente
  6. Erros de API vão para 'erros_importacao.txt'

INSTRUÇÕES:
  1. Preencha JWT_TOKEN com o seu token Bearer.
  2. Coloque o caminho do XLSX em ARQUIVO_XLSX.
  3. Instale as dependências: pip install requests openpyxl
  4. Execute: python importar_clifor.py
"""

import os
import re
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

# ============================================================
# CONFIGURAÇÃO — preencha antes de rodar
# ============================================================
JWT_TOKEN   = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJzdWIiOiIxMSIsInBlcmZpbCI6IkFkbWluaXN0cmFkb3IiLCJleHAiOjE3ODI3ODE5NjUsImp0aSI6IjE0OTgyY2Q3LTk1MGMtNDc4NC04MTgyLTU2MWM4YTcxYzRjNyJ9.BdxhD9hQ6WkPw0bBIX9kCOL62xZ3KfqG6vC7w27mA0A"
BASE_URL    = "https://amsiproject-production.up.railway.app"
ARQUIVO_XLSX = r"C:\Users\schva\Downloads\AMBSI-Associados-CargaSistema.xlsx"          # caminho para o seu CSV
# ===========================================================/=

HEADERS = {
    "Authorization": f"Bearer {JWT_TOKEN}",
    "Content-Type": "application/json",
}

# Colunas esperadas no XLSX
COL_CPF          = "CPF"
COL_RG           = "RG"
COL_NOME         = "Nome Completo"
COL_NOME_USUAL   = "Nome Usual"
COL_CELULAR      = "Celular"
COL_EMAIL        = "E-mail"
COL_LOGRADOURO   = "Endereço"
COL_NUMERO       = "Número"
COL_COMPLEMENTO  = "Complemento"
COL_BAIRRO       = "Bairro"
COL_CIDADE       = "Município"
COL_UF           = "UF"
COL_CEP          = "CEP"

erros: list[str] = []


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def limpar_celula(valor) -> str:
    """Converte qualquer valor de célula para string limpa.
    - None → ''
    - Numérico → string sem casas decimais (ex: 12345678900.0 → '12345678900')
    - Quebras de linha internas → espaço
    """
    if valor is None:
        return ""
    if isinstance(valor, float) and valor == int(valor):
        valor = int(valor)
    texto = str(valor)
    texto = texto.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    return texto.strip()


def vazio(valor: str) -> bool:
    return not valor or not valor.strip()


def limpar_cpf(cpf: str) -> str:
    """Remove formatação do CPF, mantém só dígitos."""
    return re.sub(r"\D", "", cpf)


def campos_faltando(row: dict) -> list[str]:
    """Retorna lista com os nomes dos campos obrigatórios ausentes."""
    faltando = []
    if vazio(row.get(COL_CPF, "")):
        faltando.append("CPF")
    if vazio(row.get(COL_RG, "")):
        faltando.append("RG")
    if vazio(row.get(COL_NOME, "")):
        faltando.append("Nome Completo")
    tem_contato = (
        not vazio(row.get(COL_CELULAR, "")) or
        not vazio(row.get(COL_EMAIL, ""))
    )
    if not tem_contato:
        faltando.append("Celular ou E-mail")
    return faltando


def endereco_completo(row: dict) -> bool:
    campos = [COL_LOGRADOURO, COL_NUMERO, COL_BAIRRO, COL_CIDADE, COL_UF, COL_CEP]
    return all(not vazio(row.get(c, "")) for c in campos)


def registrar_erro(mensagem: str) -> None:
    print(f"  [ERRO] {mensagem}")
    erros.append(mensagem)


def ler_xlsx(caminho: str) -> tuple[list[dict], list[str]]:
    """Lê o XLSX e retorna (lista de dicts por linha, lista de colunas)."""
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return [], []

    # Primeira linha = cabeçalho
    colunas = [limpar_celula(c) for c in linhas[0]]

    registros = []
    for linha in linhas[1:]:
        row = {colunas[i]: limpar_celula(v) for i, v in enumerate(linha)}
        # Ignora linhas completamente vazias
        if all(vazio(v) for v in row.values()):
            continue
        registros.append(row)

    return registros, colunas


# ─────────────────────────────────────────────
# Chamadas à API
# ─────────────────────────────────────────────

def criar_clifor(row: dict) -> int | None:
    """Cria o CliFor e retorna o id_clifor gerado, ou None em caso de erro."""
    cpf_limpo = limpar_cpf(row[COL_CPF])
    nome      = row[COL_NOME].strip()

    payload = {
        "pessoafisica_juridica": True,
        "cpf_cnpj":              cpf_limpo,
        "rg_inscricaoestadual":  row[COL_RG].strip(),
        "nome":                  nome,
        "tipo_clifor":           "C",
        "ativo":                 True,
        "inadimplente":          False,
    }

    try:
        resp = requests.post(
            f"{BASE_URL}/cliente_fornecedor/",
            json=payload,
            headers=HEADERS,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            data = resp.json()
            return data.get("id_clifor")
        else:
            registrar_erro(
                f"CliFor '{nome}' (CPF {cpf_limpo}) — "
                f"HTTP {resp.status_code}: {resp.text[:300]}"
            )
            return None
    except Exception as exc:
        registrar_erro(f"CliFor '{nome}' — Exceção: {exc}")
        return None


def criar_contato(id_clifor: int, tipo: str, info: str, principal: bool) -> None:
    payload = {
        "id_clifor_fk":      id_clifor,
        "tipocontato":       tipo,
        "info_do_contato":   info.strip(),
        "contato_principal": principal,
    }
    try:
        resp = requests.post(
            f"{BASE_URL}/contato/",
            json=payload,
            headers=HEADERS,
            timeout=30,
        )
        if resp.status_code not in (200, 201):
            registrar_erro(
                f"Contato {tipo} para clifor {id_clifor} — "
                f"HTTP {resp.status_code}: {resp.text[:300]}"
            )
        else:
            print(f"    [OK] Contato {tipo} cadastrado.")
    except Exception as exc:
        registrar_erro(f"Contato {tipo} para clifor {id_clifor} — Exceção: {exc}")


def criar_endereco(id_clifor: int, row: dict) -> None:
    payload = {
        "id_clifor_fk":      id_clifor,
        "enderecoprimario":  True,
        "logradouro":        row[COL_LOGRADOURO].strip(),
        "numero":            row[COL_NUMERO].strip(),
        "complemento":       row.get(COL_COMPLEMENTO, "").strip() or None,
        "bairro":            row[COL_BAIRRO].strip(),
        "cidade":            row[COL_CIDADE].strip(),
        "uf":                row[COL_UF].strip().upper(),
        "cep":               row[COL_CEP].strip(),
    }
    try:
        resp = requests.post(
            f"{BASE_URL}/endereco/",
            json=payload,
            headers=HEADERS,
            timeout=30,
        )
        if resp.status_code not in (200, 201):
            registrar_erro(
                f"Endereço para clifor {id_clifor} — "
                f"HTTP {resp.status_code}: {resp.text[:300]}"
            )
        else:
            print(f"    [OK] Endereço cadastrado.")
    except Exception as exc:
        registrar_erro(f"Endereço para clifor {id_clifor} — Exceção: {exc}")


# ─────────────────────────────────────────────
# Salvar arquivos de saída
# ─────────────────────────────────────────────

def salvar_pendentes(pendentes: list[tuple[dict, list[str]]], colunas: list[str]) -> None:
    """Gera um .xlsx com os registros pendentes e uma coluna 'Campos Faltando'."""
    if not pendentes:
        print("\nNenhum registro pendente.")
        return

    base    = os.path.splitext(os.path.basename(ARQUIVO_XLSX))[0]
    arquivo = f"pendentes_{base}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Pendentes"

    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    falta_fill  = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    cabecalho = list(colunas) + ["Campos Faltando"]
    for col_idx, titulo in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = header_fill
        cell.font = header_font

    col_faltando = len(cabecalho)

    for row_idx, (row, faltando) in enumerate(pendentes, start=2):
        for col_idx, col in enumerate(colunas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col, ""))

        cell_falta = ws.cell(
            row=row_idx,
            column=col_faltando,
            value=", ".join(faltando),
        )
        cell_falta.fill = falta_fill
        cell_falta.font = Font(bold=True, color="C00000")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(arquivo)
    print(f"\n{len(pendentes)} registro(s) pendente(s) salvo(s) em '{arquivo}'.")


def salvar_erros() -> None:
    if not erros:
        print("Nenhum erro de API registrado.")
        return

    arquivo = "erros_importacao.txt"
    with open(arquivo, "w", encoding="utf-8") as f:
        f.write(f"Total de erros: {len(erros)}\n")
        f.write("=" * 55 + "\n\n")
        for linha in erros:
            f.write(linha + "\n")

    print(f"{len(erros)} erro(s) de API registrado(s) em '{arquivo}'.")


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main() -> None:
    if JWT_TOKEN == "SEU_TOKEN_AQUI":
        print("ATENÇÃO: Preencha a variável JWT_TOKEN antes de executar o script.")
        return

    if not os.path.exists(ARQUIVO_XLSX):
        print(f"ATENÇÃO: Arquivo '{ARQUIVO_XLSX}' não encontrado.")
        return

    print(f"Lendo '{ARQUIVO_XLSX}'...")

    todas, colunas = ler_xlsx(ARQUIVO_XLSX)

    if not todas:
        print("Nenhuma linha encontrada no arquivo.")
        return

    print(f"{len(todas)} linha(s) encontrada(s).\n")

    pendentes: list[tuple[dict, list[str]]] = []
    cadastrados = 0

    for i, row in enumerate(todas, start=1):
        nome_log = row.get(COL_NOME, f"linha {i}").strip() or f"linha {i}"
        print(f"[{i}/{len(todas)}] {nome_log}")

        faltando = campos_faltando(row)
        if faltando:
            print(f"  → Incompleto. Faltando: {', '.join(faltando)}. Enviando para pendentes.")
            pendentes.append((row, faltando))
            continue

        # 1. Criar CliFor
        id_clifor = criar_clifor(row)
        if id_clifor is None:
            print("  → Falha na API ao criar CliFor (ver erros_importacao.txt).")
            continue

        print(f"  [OK] CliFor criado (id={id_clifor}).")

        # 2. Criar contato(s)
        tem_celular = not vazio(row.get(COL_CELULAR, ""))
        tem_email   = not vazio(row.get(COL_EMAIL, ""))

        if tem_celular:
            criar_contato(id_clifor, "Celular", row[COL_CELULAR], principal=not tem_email)

        if tem_email:
            criar_contato(id_clifor, "E-mail", row[COL_EMAIL], principal=True)

        # 3. Criar endereço (se completo)
        if endereco_completo(row):
            criar_endereco(id_clifor, row)
        else:
            print("  → Endereço incompleto, pulando cadastro de endereço.")

        cadastrados += 1

    print(f"\n{'='*55}")
    print(f"Concluído: {cadastrados} cadastrado(s), {len(pendentes)} pendente(s).")

    salvar_pendentes(pendentes, list(colunas))
    salvar_erros()


if __name__ == "__main__":
    main()

# ============================================================
# CONFIGURAÇÃO — preencha antes de rodar
# ============================================================
JWT_TOKEN   = "SEU_TOKEN_AQUI"
BASE_URL    = "https://amsiproject-production.up.railway.app"
ARQUIVO_CSV = "clifor.csv"          # caminho para o seu CSV
# ============================================================

HEADERS = {
    "Authorization": f"Bearer {JWT_TOKEN}",
    "Content-Type": "application/json",
}

# Colunas esperadas no CSV
COL_CPF          = "CPF"
COL_RG           = "RG"
COL_NOME         = "Nome Completo"
COL_NOME_USUAL   = "Nome Usual"
COL_CELULAR      = "Celular"
COL_EMAIL        = "E-mail"
COL_LOGRADOURO   = "Endereço"
COL_NUMERO       = "Número"
COL_COMPLEMENTO  = "Complemento"
COL_BAIRRO       = "Bairro"
COL_CIDADE       = "Município"
COL_UF           = "UF"
COL_CEP          = "CEP"

erros: list[str] = []


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def vazio(valor: str) -> bool:
    return not valor or not valor.strip()


def limpar_cpf(cpf: str) -> str:
    """Remove formatação do CPF, mantém só dígitos."""
    return re.sub(r"\D", "", cpf)


def campos_faltando(row: dict) -> list[str]:
    """Retorna lista com os nomes dos campos obrigatórios ausentes."""
    faltando = []
    if vazio(row.get(COL_CPF, "")):
        faltando.append("CPF")
    if vazio(row.get(COL_RG, "")):
        faltando.append("RG")
    if vazio(row.get(COL_NOME, "")):
        faltando.append("Nome Completo")
    tem_contato = (
        not vazio(row.get(COL_CELULAR, "")) or
        not vazio(row.get(COL_EMAIL, ""))
    )
    if not tem_contato:
        faltando.append("Celular ou E-mail")
    return faltando


def completo(row: dict) -> bool:
    """Retorna True se o registro tem os campos mínimos para cadastro."""
    return len(campos_faltando(row)) == 0


def endereco_completo(row: dict) -> bool:
    campos = [COL_LOGRADOURO, COL_NUMERO, COL_BAIRRO, COL_CIDADE, COL_UF, COL_CEP]
    return all(not vazio(row.get(c, "")) for c in campos)


def registrar_erro(mensagem: str) -> None:
    print(f"  [ERRO] {mensagem}")
    erros.append(mensagem)


# ─────────────────────────────────────────────
# Chamadas à API
# ─────────────────────────────────────────────

def criar_clifor(row: dict) -> int | None:
    """Cria o CliFor e retorna o id_clifor gerado, ou None em caso de erro."""
    cpf_limpo = limpar_cpf(row[COL_CPF])
    nome      = row[COL_NOME].strip()

    payload = {
        "pessoafisica_juridica": True,
        "cpf_cnpj":              cpf_limpo,
        "rg_inscricaoestadual":  row[COL_RG].strip(),
        "nome":                  nome,
        "tipo_clifor":           "C",
        "ativo":                 True,
        "inadimplente":          False,
    }

    # DataNascimento é opcional — omite se ausente
    # (campo não existe no CSV; deixamos sem enviar)

    try:
        resp = requests.post(
            f"{BASE_URL}/cliente_fornecedor/",
            json=payload,
            headers=HEADERS,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            data = resp.json()
            return data.get("id_clifor")
        else:
            registrar_erro(
                f"CliFor '{nome}' (CPF {cpf_limpo}) — "
                f"HTTP {resp.status_code}: {resp.text[:300]}"
            )
            return None
    except Exception as exc:
        registrar_erro(f"CliFor '{nome}' — Exceção: {exc}")
        return None


def criar_contato(id_clifor: int, tipo: str, info: str, principal: bool) -> None:
    payload = {
        "id_clifor_fk":      id_clifor,
        "tipocontato":       tipo,
        "info_do_contato":   info.strip(),
        "contato_principal": principal,
    }
    try:
        resp = requests.post(
            f"{BASE_URL}/contato/",
            json=payload,
            headers=HEADERS,
            timeout=30,
        )
        if resp.status_code not in (200, 201):
            registrar_erro(
                f"Contato {tipo} para clifor {id_clifor} — "
                f"HTTP {resp.status_code}: {resp.text[:300]}"
            )
        else:
            print(f"    [OK] Contato {tipo} cadastrado.")
    except Exception as exc:
        registrar_erro(f"Contato {tipo} para clifor {id_clifor} — Exceção: {exc}")


def criar_endereco(id_clifor: int, row: dict) -> None:
    payload = {
        "id_clifor_fk":      id_clifor,
        "enderecoprimario":  True,
        "logradouro":        row[COL_LOGRADOURO].strip(),
        "numero":            row[COL_NUMERO].strip(),
        "complemento":       row.get(COL_COMPLEMENTO, "").strip() or None,
        "bairro":            row[COL_BAIRRO].strip(),
        "cidade":            row[COL_CIDADE].strip(),
        "uf":                row[COL_UF].strip().upper(),
        "cep":               row[COL_CEP].strip(),
    }
    try:
        resp = requests.post(
            f"{BASE_URL}/endereco/",
            json=payload,
            headers=HEADERS,
            timeout=30,
        )
        if resp.status_code not in (200, 201):
            registrar_erro(
                f"Endereço para clifor {id_clifor} — "
                f"HTTP {resp.status_code}: {resp.text[:300]}"
            )
        else:
            print(f"    [OK] Endereço cadastrado.")
    except Exception as exc:
        registrar_erro(f"Endereço para clifor {id_clifor} — Exceção: {exc}")


# ─────────────────────────────────────────────
# Salvar arquivos de saída
# ─────────────────────────────────────────────

def salvar_pendentes(pendentes: list[tuple[dict, list[str]]], colunas: list[str]) -> None:
    """Gera um .xlsx com os registros pendentes e uma coluna 'Campos Faltando'."""
    if not pendentes:
        print("\nNenhum registro pendente.")
        return

    base    = os.path.splitext(os.path.basename(ARQUIVO_CSV))[0]
    arquivo = f"pendentes_{base}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Pendentes"

    # Estilos
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    falta_fill  = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Cabeçalho — colunas originais + coluna extra
    cabecalho = list(colunas) + ["Campos Faltando"]
    for col_idx, titulo in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = header_fill
        cell.font = header_font

    col_faltando = len(cabecalho)  # índice da coluna extra (1-based)

    # Dados
    for row_idx, (row, faltando) in enumerate(pendentes, start=2):
        for col_idx, col in enumerate(colunas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col, ""))

        # Coluna "Campos Faltando" com destaque
        cell_falta = ws.cell(
            row=row_idx,
            column=col_faltando,
            value=", ".join(faltando),
        )
        cell_falta.fill = falta_fill
        cell_falta.font = Font(bold=True, color="C00000")

    # Ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(arquivo)
    print(f"\n{len(pendentes)} registro(s) pendente(s) salvo(s) em '{arquivo}'.")


def salvar_erros() -> None:
    if not erros:
        print("Nenhum erro de API registrado.")
        return

    arquivo = "erros_importacao.txt"
    with open(arquivo, "w", encoding="utf-8") as f:
        f.write(f"Total de erros: {len(erros)}\n")
        f.write("=" * 55 + "\n\n")
        for linha in erros:
            f.write(linha + "\n")

    print(f"{len(erros)} erro(s) de API registrado(s) em '{arquivo}'.")


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main() -> None:
    if JWT_TOKEN == "SEU_TOKEN_AQUI":
        print("ATENÇÃO: Preencha a variável JWT_TOKEN antes de executar o script.")
        return

    if not os.path.exists(ARQUIVO_CSV):
        print(f"ATENÇÃO: Arquivo '{ARQUIVO_CSV}' não encontrado.")
        return

    print(f"Lendo '{ARQUIVO_CSV}'...")

    with open(ARQUIVO_CSV, newline="", encoding="utf-8-sig") as f:
        reader    = csv.DictReader(f, delimiter=",")
        todas     = list(reader)
        colunas   = reader.fieldnames or []

    print(f"{len(todas)} linha(s) encontrada(s).\n")

    pendentes: list[tuple[dict, list[str]]] = []
    cadastrados = 0

    for i, row in enumerate(todas, start=1):
        nome_log = row.get(COL_NOME, f"linha {i}").strip() or f"linha {i}"
        print(f"[{i}/{len(todas)}] {nome_log}")

        faltando = campos_faltando(row)
        if faltando:
            print(f"  → Incompleto. Faltando: {', '.join(faltando)}. Enviando para pendentes.")
            pendentes.append((row, faltando))
            continue

        # 1. Criar CliFor
        id_clifor = criar_clifor(row)
        if id_clifor is None:
            # Erro já registrado em erros_importacao.txt pela função
            print("  → Falha na API ao criar CliFor (ver erros_importacao.txt).")
            continue

        print(f"  [OK] CliFor criado (id={id_clifor}).")

        # 2. Criar contato(s)
        tem_celular = not vazio(row.get(COL_CELULAR, ""))
        tem_email   = not vazio(row.get(COL_EMAIL, ""))

        if tem_celular:
            # Celular é o principal se não houver e-mail
            criar_contato(id_clifor, "Celular", row[COL_CELULAR], principal=not tem_email)

        if tem_email:
            # E-mail é sempre principal (ou único)
            criar_contato(id_clifor, "E-mail", row[COL_EMAIL], principal=True)

        # 3. Criar endereço (se completo)
        if endereco_completo(row):
            criar_endereco(id_clifor, row)
        else:
            print("  → Endereço incompleto, pulando cadastro de endereço.")

        cadastrados += 1

    # Relatórios finais
    print(f"\n{'='*55}")
    print(f"Concluído: {cadastrados} cadastrado(s), {len(pendentes)} pendente(s).")

    salvar_pendentes(pendentes, list(colunas))
    salvar_erros()


if __name__ == "__main__":
    main()